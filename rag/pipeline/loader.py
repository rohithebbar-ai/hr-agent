# rag/pipeline/loader.py

"""
Document Loader — Three-Tier PDF Fallback
──────────────────────────────────────────
PDF routing:
  Tier 1: LlamaParse    — best quality, costs credits
  Tier 2: Unstructured  — free fallback, augmented with PyMuPDF tables
  Tier 3: PyMuPDF       — last resort, always works

Non-PDF routing:
  DOCX/DOC → Unstructured
  TXT/MD   → Direct read + markdown parser
  XLSX/XLS → openpyxl
  JSON     → Custom flattener

All formats return LoadedDocument with List[DocumentElement].
Everything downstream (chunker, embedder, upserter) is format-agnostic.
"""

import hashlib
import os
import re
import requests
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple
from dotenv import load_dotenv
load_dotenv()

# Shared HTTP session — reuses TCP connections across LlamaParse/Qdrant health calls
_http_session = requests.Session()


# ── Credit exhaustion error signatures from LlamaParse ──
QUOTA_ERROR_SIGNATURES = [
    "quota",
    "credit",
    "limit exceeded",
    "insufficient",
    "402",
    "payment required",
    "out of credits",
]


@dataclass
class DocumentElement:
    """
    A single typed content block.

    element_type:
        Title         → section heading (sets section_path in chunker)
        NarrativeText → body paragraph
        Table         → structured table in markdown format
        ListItem      → bullet or numbered list item
        Header        → page header   (chunker discards)
        Footer        → page footer   (chunker discards)
    """
    element_type: str
    text: str
    metadata: dict = field(default_factory=dict)
    html: Optional[str] = None   # markdown table text stored here too


@dataclass
class LoadedDocument:
    document_id: str        # SHA256 — deterministic deduplication
    filename: str
    file_type: str
    title: str
    page_count: int
    elements: List[DocumentElement]
    file_size_bytes: int
    loader_used: str        # "llamaparse" | "unstructured" | "pymupdf" | "direct"


def load_document(
    file_path: Path,
    filename: str,
    use_llamaparse: bool = True,
) -> LoadedDocument:
    """
    Load any document and return typed elements.

    Args:
        file_path:      Path to the file on disk
        filename:       Original filename (used for extension detection)
        use_llamaparse: If False, skip LlamaParse and use Unstructured.
                        Admin can set this per-document to save credits.
    """
    raw_bytes = file_path.read_bytes()
    document_id = hashlib.sha256(raw_bytes).hexdigest()[:16]
    suffix = Path(filename).suffix.lower()
    file_size = len(raw_bytes)

    print(f"[LOADER] {filename} ({file_size // 1024} KB) — format: {suffix}")

    if suffix == ".pdf":
        elements, loader_used = _load_pdf(file_path, use_llamaparse)
    elif suffix in (".docx", ".doc"):
        elements = _load_docx_unstructured(file_path)
        loader_used = "unstructured"
    elif suffix in (".txt", ".md"):
        elements = _load_text(file_path, suffix)
        loader_used = "direct"
    elif suffix in (".xlsx", ".xls"):
        elements = _load_excel(file_path)
        loader_used = "openpyxl"
    elif suffix == ".json":
        elements = _load_json(file_path)
        loader_used = "json_flattener"
    else:
        raise ValueError(
            f"Unsupported file type: {suffix}. "
            f"Supported: pdf, docx, doc, txt, md, xlsx, xls, json"
        )

    # Extract title from first Title element
    title = filename
    for el in elements:
        if el.element_type == "Title" and len(el.text) > 3:
            title = el.text
            break

    # Estimate page count
    page_numbers = [
        el.metadata.get("page_number", 0)
        for el in elements
        if el.metadata.get("page_number")
    ]
    page_count = max(page_numbers) if page_numbers else 1

    table_count = sum(1 for el in elements if el.element_type == "Table")
    print(
        f"[LOADER] Done via {loader_used} — "
        f"{len(elements)} elements, {page_count} pages, {table_count} tables"
    )

    return LoadedDocument(
        document_id=document_id,
        filename=filename,
        file_type=suffix.lstrip("."),
        title=title,
        page_count=page_count,
        elements=elements,
        file_size_bytes=file_size,
        loader_used=loader_used,
    )


# ─────────────────────────────────────────────────────────────────
# PDF: Three-tier fallback
# ─────────────────────────────────────────────────────────────────

def _load_pdf(
    file_path: Path,
    use_llamaparse: bool,
) -> Tuple[List[DocumentElement], str]:
    """
    Three-tier PDF loading with automatic fallback.
    Returns (elements, loader_name).
    """

    # ── Tier 1: LlamaParse ──
    if use_llamaparse:
        api_key = os.environ.get("LLAMAPARSE_API_KEY")
        if not api_key:
            print("[LOADER] LLAMAPARSE_API_KEY not set — skipping to Unstructured")
        else:
            elements, success, is_quota_error = _try_llamaparse(file_path, api_key)
            if success:
                return elements, "llamaparse"

            if is_quota_error:
                print("[LOADER] LlamaParse quota exhausted — switching to Unstructured + PyMuPDF")
                print("[LOADER]Consider upgrading at cloud.llamaindex.ai")
            else:
                print("[LOADER] LlamaParse failed — falling back to Unstructured + PyMuPDF")
    else:
        print("[LOADER] LlamaParse skipped by admin — using Unstructured + PyMuPDF")

    # ── Tier 2: Unstructured fast + PyMuPDF table augmentation ──
    elements, success = _try_unstructured_with_pymupdf(file_path)
    if success:
        return elements, "unstructured+pymupdf"

    # ── Tier 3: PyMuPDF direct (last resort) ──
    print("[LOADER] Unstructured failed — using PyMuPDF direct extraction")
    elements = _load_pdf_pymupdf_direct(file_path)
    return elements, "pymupdf"


def _try_llamaparse(
    file_path: Path,
    api_key: str,
) -> Tuple[List[DocumentElement], bool, bool]:
    """
    Attempt LlamaParse.
    Returns (elements, success, is_quota_error).
    """
    try:
        from llama_parse import LlamaParse

        parser = LlamaParse(
            api_key=api_key,
            result_type="markdown",
            verbose=False,
            language="en",
            skip_diagonal_text=True,
        )

        print("[LOADER] Calling LlamaParse API")
        documents = parser.load_data(str(file_path))

        if not documents:
            print("[LOADER] LlamaParse returned empty result")
            return [], False, False

        all_elements = []
        for page_idx, doc in enumerate(documents):
            page_elements = _parse_markdown_to_elements(
                text=doc.text,
                page_number=page_idx + 1,
            )
            all_elements.extend(page_elements)

        if not all_elements:
            return [], False, False

        return all_elements, True, False

    except Exception as e:
        error_str = str(e).lower()
        is_quota = any(sig in error_str for sig in QUOTA_ERROR_SIGNATURES)

        if is_quota:
            print(f"[LOADER] LlamaParse quota error: {e}")
        else:
            print(f"[LOADER] LlamaParse error: {type(e).__name__}: {e}")

        return [], False, is_quota


def _try_unstructured_with_pymupdf(
    file_path: Path,
) -> Tuple[List[DocumentElement], bool]:
    """
    Tier 2: Unstructured fast + PyMuPDF table detection.
    Unstructured classifies elements. PyMuPDF catches tables Unstructured misses.
    Returns (elements, success).
    """
    try:
        from unstructured.partition.auto import partition

        print("[LOADER] Using Unstructured fast strategy")
        raw_elements = partition(
            filename=str(file_path),
            strategy="fast",
            include_page_breaks=True,
            extract_images_in_pdf=False,
            infer_table_structure=True,
        )

        elements = []
        for el in raw_elements:
            element_type = type(el).__name__
            text = str(el).strip()
            if not text or element_type == "PageBreak":
                continue

            html = None
            if element_type == "Table" and hasattr(el, "metadata"):
                html = getattr(el.metadata, "text_as_html", None)

            metadata = {}
            if hasattr(el, "metadata"):
                metadata = {
                    "page_number": getattr(el.metadata, "page_number", None),
                }

            elements.append(DocumentElement(
                element_type=element_type,
                text=text,
                metadata=metadata,
                html=html,
            ))

        if not elements:
            return [], False

        # Augment with PyMuPDF table detection
        elements = _augment_with_pymupdf_tables(file_path, elements)

        return elements, True

    except Exception as e:
        print(f"[LOADER] Unstructured failed: {type(e).__name__}: {e}")
        return [], False


def _augment_with_pymupdf_tables(
    file_path: Path,
    elements: List[DocumentElement],
) -> List[DocumentElement]:
    """
    Detect tables using PyMuPDF that Unstructured fast missed.
    Only runs if Unstructured found 0 tables — avoids duplicates.
    """
    existing_tables = sum(1 for el in elements if el.element_type == "Table")
    if existing_tables > 0:
        print(f"[LOADER] Unstructured already found {existing_tables} tables — skipping PyMuPDF augmentation")
        return elements

    try:
        import fitz

        doc = fitz.open(str(file_path))
        table_elements = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            try:
                tabs = page.find_tables()
                for tab in tabs.tables:
                    table_md = _pymupdf_table_to_markdown(tab)
                    if not table_md.strip():
                        continue
                    table_elements.append(DocumentElement(
                        element_type="Table",
                        text=table_md,
                        metadata={"page_number": page_num + 1},
                        html=table_md,
                    ))
            except Exception:
                continue

        if table_elements:
            print(f"[LOADER] PyMuPDF found {len(table_elements)} tables Unstructured missed")
            elements = _insert_table_elements(elements, table_elements)
        else:
            print("[LOADER] PyMuPDF found no additional tables")

        return elements

    except Exception as e:
        print(f"[LOADER] PyMuPDF augmentation failed: {e}")
        return elements


def _pymupdf_table_to_markdown(tab) -> str:
    """Convert PyMuPDF table object to markdown."""
    try:
        rows = tab.extract()
        if not rows:
            return ""
        rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows]
        if not rows:
            return ""
        header = rows[0]
        separator = ["---"] * len(header)
        lines = [" | ".join(header), " | ".join(separator)]
        for row in rows[1:]:
            lines.append(" | ".join(row))
        return "\n".join(lines)
    except Exception:
        return ""


def _insert_table_elements(
    elements: List[DocumentElement],
    table_elements: List[DocumentElement],
) -> List[DocumentElement]:
    """Insert table elements at the correct position by page number."""
    combined = list(elements)
    for table_el in table_elements:
        table_page = table_el.metadata.get("page_number", 1)
        insert_pos = len(combined)
        for i, el in enumerate(combined):
            if el.metadata.get("page_number", 0) > table_page:
                insert_pos = i
                break
        combined.insert(insert_pos, table_el)
    return combined


def _load_pdf_pymupdf_direct(file_path: Path) -> List[DocumentElement]:
    """
    Tier 3: PyMuPDF direct extraction — last resort.
    No element classification. Plain text per page.
    Always works even on unusual PDFs.
    """
    try:
        import fitz
        doc = fitz.open(str(file_path))
        elements = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text").strip()
            if not text:
                continue

            # Basic heading detection from font sizes
            page_elements = _pymupdf_classify_page(page, page_num + 1)
            elements.extend(page_elements)

        print(f"[LOADER] PyMuPDF direct: {len(elements)} elements")
        return elements

    except Exception as e:
        print(f"[LOADER] PyMuPDF direct failed: {e}")
        return []


def _pymupdf_classify_page(page, page_number: int) -> List[DocumentElement]:
    """
    Classify text blocks by font size to detect headings.
    Large font = Title, normal font = NarrativeText.
    """
    import fitz

    blocks = page.get_text("dict")["blocks"]
    elements = []

    for block in blocks:
        if block.get("type") != 0:  # 0 = text block
            continue

        block_text = ""
        max_font_size = 0

        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = span.get("text", "").strip()
                if text:
                    block_text += text + " "
                    size = span.get("size", 0)
                    if size > max_font_size:
                        max_font_size = size

        block_text = block_text.strip()
        if not block_text:
            continue

        # Classify by font size
        # > 14pt = likely heading, <= 14pt = body text
        element_type = "Title" if max_font_size > 14 else "NarrativeText"

        elements.append(DocumentElement(
            element_type=element_type,
            text=block_text,
            metadata={"page_number": page_number, "font_size": max_font_size},
        ))

    return elements


# ─────────────────────────────────────────────────────────────────
# Markdown parser (shared by LlamaParse PDF output and .md files)
# ─────────────────────────────────────────────────────────────────

def _parse_markdown_to_elements(
    text: str,
    page_number: int,
) -> List[DocumentElement]:
    """
    Convert markdown text to DocumentElement list.

    LlamaParse returns markdown — this classifies each block:
    # ## ###       → Title (with heading level in metadata)
    | col | col |  → Table
    - item / 1.    → ListItem
    plain text     → NarrativeText
    ---            → skip (horizontal rule)
    """
    elements = []
    lines = text.split("\n")
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # Skip empty lines and horizontal rules
        if not stripped or stripped in ("---", "***", "___"):
            i += 1
            continue

        # Heading
        heading_match = re.match(r'^(#{1,4})\s+(.+)$', stripped)
        if heading_match:
            elements.append(DocumentElement(
                element_type="Title",
                text=heading_match.group(2).strip(),
                metadata={
                    "page_number": page_number,
                    "heading_level": len(heading_match.group(1)),
                },
            ))
            i += 1
            continue

        # Table (lines starting with |)
        if stripped.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i].strip())
                i += 1

            # Remove separator row (| --- | --- |)
            table_rows = [
                row for row in table_lines
                if not re.match(r'^\|[\s\-|:]+\|$', row)
            ]
            if table_rows:
                table_text = "\n".join(table_rows)
                elements.append(DocumentElement(
                    element_type="Table",
                    text=table_text,
                    metadata={"page_number": page_number},
                    html=table_text,
                ))
            continue

        # List item (-, *, •, 1.)
        if re.match(r'^[-*•]\s+', stripped) or re.match(r'^\d+\.\s+', stripped):
            list_text = re.sub(r'^[-*•]\s+', '', stripped)
            list_text = re.sub(r'^\d+\.\s+', '', list_text)
            elements.append(DocumentElement(
                element_type="ListItem",
                text=list_text,
                metadata={"page_number": page_number},
            ))
            i += 1
            continue

        # NarrativeText — collect consecutive plain lines
        narrative_lines = []
        while i < len(lines):
            stripped_line = lines[i].strip()
            if (not stripped_line
                    or stripped_line.startswith("#")
                    or stripped_line.startswith("|")
                    or re.match(r'^[-*•]\s+', stripped_line)
                    or re.match(r'^\d+\.\s+', stripped_line)
                    or stripped_line in ("---", "***", "___")):
                break
            narrative_lines.append(stripped_line)
            i += 1

        if narrative_lines:
            narrative_text = " ".join(narrative_lines)
            if len(narrative_text) > 10:
                elements.append(DocumentElement(
                    element_type="NarrativeText",
                    text=narrative_text,
                    metadata={"page_number": page_number},
                ))

    return elements


# ─────────────────────────────────────────────────────────────────
# DOCX: Unstructured
# ─────────────────────────────────────────────────────────────────

def _load_docx_unstructured(file_path: Path) -> List[DocumentElement]:
    """Parse DOCX with Unstructured. Handles Word styles and heading levels."""
    from unstructured.partition.auto import partition

    raw_elements = partition(
        filename=str(file_path),
        strategy="fast",
        include_page_breaks=False,
    )

    elements = []
    for el in raw_elements:
        element_type = type(el).__name__
        text = str(el).strip()
        if not text or element_type == "PageBreak":
            continue

        html = None
        if element_type == "Table" and hasattr(el, "metadata"):
            html = getattr(el.metadata, "text_as_html", None)

        metadata = {}
        if hasattr(el, "metadata"):
            metadata = {"page_number": getattr(el.metadata, "page_number", 1)}

        elements.append(DocumentElement(
            element_type=element_type,
            text=text,
            metadata=metadata,
            html=html,
        ))

    return elements


# ─────────────────────────────────────────────────────────────────
# TXT / MD: Direct read
# ─────────────────────────────────────────────────────────────────

def _load_text(file_path: Path, suffix: str) -> List[DocumentElement]:
    """Load plain text or markdown. MD gets parsed into typed elements."""
    content = file_path.read_text(encoding="utf-8", errors="replace")

    if suffix == ".md":
        return _parse_markdown_to_elements(content, page_number=1)
    else:
        paragraphs = [p.strip() for p in content.split("\n\n") if p.strip()]
        return [
            DocumentElement(
                element_type="NarrativeText",
                text=para,
                metadata={"page_number": 1},
            )
            for para in paragraphs
            if len(para) > 10
        ]


# ─────────────────────────────────────────────────────────────────
# XLSX: openpyxl
# ─────────────────────────────────────────────────────────────────

def _load_excel(file_path: Path) -> List[DocumentElement]:
    """Load Excel — one Table element per sheet with markdown formatting."""
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    elements = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        elements.append(DocumentElement(
            element_type="Title",
            text=sheet_name,
            metadata={"page_number": 1},
        ))

        if len(rows) >= 2:
            header = rows[0]
            sep = ["---"] * len(header)
            table_md = "\n".join([
                " | ".join(header),
                " | ".join(sep),
                *[" | ".join(row) for row in rows[1:]],
            ])
        else:
            table_md = " | ".join(rows[0])

        elements.append(DocumentElement(
            element_type="Table",
            text=table_md,
            metadata={"page_number": 1, "sheet_name": sheet_name},
            html=table_md,
        ))

    return elements

# ─────────────────────────────────────────────────────────────────
# JSON: Custom flattener
# ─────────────────────────────────────────────────────────────────

def _load_json(file_path: Path) -> List[DocumentElement]:
    """Flatten JSON to readable text. Each top-level key = section."""
    import json

    data = json.loads(file_path.read_text(encoding="utf-8"))
    elements = []

    def flatten(obj, prefix=""):
        if isinstance(obj, dict):
            for key, value in obj.items():
                section_key = f"{prefix} > {key}" if prefix else key
                if isinstance(value, (dict, list)):
                    elements.append(DocumentElement(
                        element_type="Title",
                        text=section_key,
                        metadata={"page_number": 1},
                    ))
                    flatten(value, section_key)
                else:
                    elements.append(DocumentElement(
                        element_type="NarrativeText",
                        text=f"{key}: {value}",
                        metadata={"page_number": 1},
                    ))
        elif isinstance(obj, list):
            for item in obj:
                if isinstance(item, (dict, list)):
                    flatten(item, prefix)
                else:
                    elements.append(DocumentElement(
                        element_type="ListItem",
                        text=str(item),
                        metadata={"page_number": 1},
                    ))

    flatten(data)
    return elements


# ─────────────────────────────────────────────────────────────────
# CLEANER: Custom cleaner function to remove header and footer.
# ─────────────────────────────────────────────────────────────────

def clean_elements(elements: List[DocumentElement]) -> List[DocumentElement]:
    """
    Remove noise elements that hurt retrieval quality.
    Called after loading, before chunking.
    """
    from collections import Counter
    
    # Count how many times each short text appears
    # Repeated short text = header/footer
    short_texts = [
        el.text.strip()
        for el in elements
        if len(el.text.strip()) < 80
        and el.element_type not in ("Table",)
    ]
    text_counts = Counter(short_texts)
    
    # Text appearing 3+ times on different pages = likely header/footer
    repeated = {
        text for text, count in text_counts.items()
        if count >= 3 and len(text) > 5
    }
    
    cleaned = []
    for el in elements:
        # Always discard Header/Footer from Unstructured
        if el.element_type in ("Header", "Footer"):
            continue
        
        # Discard repeated short text (page headers/footers in LlamaParse)
        if el.text.strip() in repeated and el.element_type != "Table":
            continue
        
        # Discard very short non-title elements (page numbers, "—", etc.)
        if len(el.text.strip()) < 10 and el.element_type != "Title":
            continue
        
        cleaned.append(el)
    
    removed = len(elements) - len(cleaned)
    if removed > 0:
        print(f"[CLEANER] Removed {removed} noise elements")
    
    return cleaned